"""
Report Generator - Generate PDF and Excel trading reports
Supports customizable templates and scheduled generation
"""
import os
import io
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Any
from pathlib import Path

from src.core.logger import get_logger
from src.core.statistics import TradingStatistics, PerformanceMetrics

logger = get_logger(__name__)


@dataclass
class ReportConfig:
    """Report configuration"""
    title: str = "AI Futu Trader Report"
    subtitle: str = "Trading Performance Report"
    author: str = "AI Futu Trader"

    # Sections to include
    include_summary: bool = True
    include_performance: bool = True
    include_trades: bool = True
    include_positions: bool = True
    include_charts: bool = True
    include_risk: bool = True

    # Style
    primary_color: str = "#1E88E5"
    secondary_color: str = "#43A047"
    danger_color: str = "#E53935"

    # Output
    output_dir: str = "reports"


@dataclass
class ReportData:
    """Data for report generation"""
    start_date: date
    end_date: date

    # Summary
    starting_equity: float = 0.0
    ending_equity: float = 0.0
    total_pnl: float = 0.0
    total_pnl_pct: float = 0.0

    # Performance
    metrics: Optional[PerformanceMetrics] = None

    # Daily data
    daily_pnl: List[Dict] = field(default_factory=list)
    equity_curve: List[Dict] = field(default_factory=list)

    # Trades
    trades: List[Dict] = field(default_factory=list)
    top_winners: List[Dict] = field(default_factory=list)
    top_losers: List[Dict] = field(default_factory=list)

    # By symbol
    symbol_performance: Dict[str, Dict] = field(default_factory=dict)


class ReportGenerator:
    """
    Generates trading reports in PDF and Excel formats.
    """

    def __init__(self, config: ReportConfig = None):
        self.config = config or ReportConfig()

        # Ensure output directory exists
        Path(self.config.output_dir).mkdir(parents=True, exist_ok=True)

    def collect_data(
        self,
        start_date: date,
        end_date: date
    ) -> ReportData:
        """Collect data for report"""
        from src.data.persistence import get_trade_database

        data = ReportData(start_date=start_date, end_date=end_date)

        try:
            db = get_trade_database()

            # Get daily performance
            daily_records = db.get_daily_performance(start_date, end_date)

            if daily_records:
                data.starting_equity = daily_records[0].starting_equity
                data.ending_equity = daily_records[-1].ending_equity
                data.total_pnl = data.ending_equity - data.starting_equity
                data.total_pnl_pct = data.total_pnl / data.starting_equity if data.starting_equity > 0 else 0

                data.daily_pnl = [
                    {
                        "date": r.date.isoformat(),
                        "pnl": r.realized_pnl,
                        "equity": r.ending_equity,
                        "trades": r.total_trades,
                        "win_rate": r.win_rate
                    }
                    for r in daily_records
                ]

                # Equity curve
                data.equity_curve = [
                    {"date": r.date.isoformat(), "equity": r.ending_equity}
                    for r in daily_records
                ]

            # Get trades
            all_trades = []
            current_date = start_date
            while current_date <= end_date:
                trades = db.get_trades_by_date(current_date)
                all_trades.extend(trades)
                current_date += timedelta(days=1)

            data.trades = [t.to_dict() for t in all_trades]

            # Top winners/losers
            sorted_trades = sorted(all_trades, key=lambda t: t.pnl, reverse=True)
            data.top_winners = [t.to_dict() for t in sorted_trades[:5] if t.pnl > 0]
            data.top_losers = [t.to_dict() for t in sorted_trades[-5:] if t.pnl < 0]

            # By symbol
            symbol_pnl = {}
            for trade in all_trades:
                if trade.symbol not in symbol_pnl:
                    symbol_pnl[trade.symbol] = {
                        "symbol": trade.symbol,
                        "trades": 0,
                        "pnl": 0,
                        "winners": 0,
                        "losers": 0
                    }
                symbol_pnl[trade.symbol]["trades"] += 1
                symbol_pnl[trade.symbol]["pnl"] += trade.pnl
                if trade.pnl > 0:
                    symbol_pnl[trade.symbol]["winners"] += 1
                else:
                    symbol_pnl[trade.symbol]["losers"] += 1

            data.symbol_performance = symbol_pnl

            # Calculate metrics
            if all_trades:
                stats = TradingStatistics(starting_equity=data.starting_equity)
                # Would need to replay trades to calculate full metrics
                # For now, use aggregate stats
                data.metrics = PerformanceMetrics(
                    total_return=data.total_pnl,
                    total_return_pct=data.total_pnl_pct,
                    total_trades=len(all_trades),
                    winning_trades=sum(1 for t in all_trades if t.pnl > 0),
                    losing_trades=sum(1 for t in all_trades if t.pnl <= 0),
                    win_rate=sum(1 for t in all_trades if t.pnl > 0) / len(all_trades) if all_trades else 0,
                    start_date=start_date,
                    end_date=end_date,
                    trading_days=len(daily_records)
                )

        except Exception as e:
            logger.error(f"Error collecting report data: {e}")

        return data

    def generate_pdf(
        self,
        start_date: date,
        end_date: date,
        filename: str = None
    ) -> str:
        """
        Generate PDF report.

        Returns:
            Path to generated PDF file
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, Image
            )
        except ImportError:
            logger.error("reportlab not installed. Install with: pip install reportlab")
            raise ImportError("reportlab required for PDF generation")

        # Collect data
        data = self.collect_data(start_date, end_date)

        # Generate filename
        if not filename:
            filename = f"trading_report_{start_date}_{end_date}.pdf"

        filepath = os.path.join(self.config.output_dir, filename)

        # Create PDF
        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )

        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor(self.config.primary_color)
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor(self.config.primary_color)
        )

        # Build content
        story = []

        # Title
        story.append(Paragraph(self.config.title, title_style))
        story.append(Paragraph(
            f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
            styles['Normal']
        ))
        story.append(Spacer(1, 30))

        # Summary Section
        if self.config.include_summary:
            story.append(Paragraph("Executive Summary", heading_style))

            pnl_color = self.config.secondary_color if data.total_pnl >= 0 else self.config.danger_color

            summary_data = [
                ["Starting Equity", f"${data.starting_equity:,.2f}"],
                ["Ending Equity", f"${data.ending_equity:,.2f}"],
                ["Total P&L", f"${data.total_pnl:+,.2f}"],
                ["Return", f"{data.total_pnl_pct*100:+.2f}%"],
            ]

            if data.metrics:
                summary_data.extend([
                    ["Total Trades", str(data.metrics.total_trades)],
                    ["Win Rate", f"{data.metrics.win_rate*100:.1f}%"],
                    ["Trading Days", str(data.metrics.trading_days)],
                ])

            summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            story.append(summary_table)
            story.append(Spacer(1, 20))

        # Performance by Symbol
        if self.config.include_performance and data.symbol_performance:
            story.append(Paragraph("Performance by Symbol", heading_style))

            symbol_data = [["Symbol", "Trades", "P&L", "Win Rate"]]
            for sym, perf in data.symbol_performance.items():
                win_rate = perf["winners"] / perf["trades"] * 100 if perf["trades"] > 0 else 0
                symbol_data.append([
                    sym,
                    str(perf["trades"]),
                    f"${perf['pnl']:+,.2f}",
                    f"{win_rate:.1f}%"
                ])

            symbol_table = Table(symbol_data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 1*inch])
            symbol_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.config.primary_color)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            story.append(symbol_table)
            story.append(Spacer(1, 20))

        # Top Trades
        if self.config.include_trades:
            if data.top_winners:
                story.append(Paragraph("Top Winning Trades", heading_style))

                winner_data = [["Symbol", "Entry", "Exit", "P&L"]]
                for trade in data.top_winners[:5]:
                    winner_data.append([
                        trade.get("symbol", ""),
                        f"${trade.get('entry_price', 0):.2f}",
                        f"${trade.get('exit_price', 0):.2f}" if trade.get('exit_price') else "-",
                        f"${trade.get('pnl', 0):+,.2f}"
                    ])

                winner_table = Table(winner_data, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1.2*inch])
                winner_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.config.secondary_color)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))

                story.append(winner_table)
                story.append(Spacer(1, 15))

            if data.top_losers:
                story.append(Paragraph("Top Losing Trades", heading_style))

                loser_data = [["Symbol", "Entry", "Exit", "P&L"]]
                for trade in data.top_losers[:5]:
                    loser_data.append([
                        trade.get("symbol", ""),
                        f"${trade.get('entry_price', 0):.2f}",
                        f"${trade.get('exit_price', 0):.2f}" if trade.get('exit_price') else "-",
                        f"${trade.get('pnl', 0):+,.2f}"
                    ])

                loser_table = Table(loser_data, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1.2*inch])
                loser_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.config.danger_color)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))

                story.append(loser_table)

        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} by {self.config.author}",
            styles['Normal']
        ))

        # Build PDF
        doc.build(story)

        logger.info(f"PDF report generated: {filepath}")
        return filepath

    def generate_excel(
        self,
        start_date: date,
        end_date: date,
        filename: str = None
    ) -> str:
        """
        Generate Excel report.

        Returns:
            Path to generated Excel file
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.chart import LineChart, Reference
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise ImportError("openpyxl required for Excel generation")

        # Collect data
        data = self.collect_data(start_date, end_date)

        # Generate filename
        if not filename:
            filename = f"trading_report_{start_date}_{end_date}.xlsx"

        filepath = os.path.join(self.config.output_dir, filename)

        # Create workbook
        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")

        # ==========================================
        # Summary Sheet
        # ==========================================
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Title
        ws_summary['A1'] = self.config.title
        ws_summary['A1'].font = Font(size=18, bold=True)
        ws_summary['A2'] = f"{start_date} to {end_date}"

        # Summary data
        summary_data = [
            ("Metric", "Value"),
            ("Starting Equity", f"${data.starting_equity:,.2f}"),
            ("Ending Equity", f"${data.ending_equity:,.2f}"),
            ("Total P&L", f"${data.total_pnl:+,.2f}"),
            ("Return %", f"{data.total_pnl_pct*100:+.2f}%"),
        ]

        if data.metrics:
            summary_data.extend([
                ("Total Trades", data.metrics.total_trades),
                ("Winning Trades", data.metrics.winning_trades),
                ("Losing Trades", data.metrics.losing_trades),
                ("Win Rate", f"{data.metrics.win_rate*100:.1f}%"),
                ("Trading Days", data.metrics.trading_days),
            ])

        for row_idx, (label, value) in enumerate(summary_data, start=4):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)

            if row_idx == 4:
                ws_summary.cell(row=row_idx, column=1).font = header_font
                ws_summary.cell(row=row_idx, column=1).fill = header_fill
                ws_summary.cell(row=row_idx, column=2).font = header_font
                ws_summary.cell(row=row_idx, column=2).fill = header_fill

        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 20

        # ==========================================
        # Daily P&L Sheet
        # ==========================================
        if data.daily_pnl:
            ws_daily = wb.create_sheet("Daily P&L")

            headers = ["Date", "P&L", "Cumulative Equity", "Trades", "Win Rate"]
            for col, header in enumerate(headers, start=1):
                cell = ws_daily.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row_idx, daily in enumerate(data.daily_pnl, start=2):
                ws_daily.cell(row=row_idx, column=1, value=daily["date"])
                ws_daily.cell(row=row_idx, column=2, value=daily["pnl"])
                ws_daily.cell(row=row_idx, column=3, value=daily["equity"])
                ws_daily.cell(row=row_idx, column=4, value=daily["trades"])
                ws_daily.cell(row=row_idx, column=5, value=daily["win_rate"])

            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_daily.column_dimensions[col].width = 15

        # ==========================================
        # Trades Sheet
        # ==========================================
        if data.trades:
            ws_trades = wb.create_sheet("Trades")

            headers = ["Trade ID", "Symbol", "Side", "Entry Time", "Entry Price",
                      "Exit Time", "Exit Price", "Quantity", "P&L", "Status"]

            for col, header in enumerate(headers, start=1):
                cell = ws_trades.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row_idx, trade in enumerate(data.trades, start=2):
                ws_trades.cell(row=row_idx, column=1, value=trade.get("trade_id", ""))
                ws_trades.cell(row=row_idx, column=2, value=trade.get("symbol", ""))
                ws_trades.cell(row=row_idx, column=3, value=trade.get("entry_side", ""))
                ws_trades.cell(row=row_idx, column=4, value=trade.get("entry_time", ""))
                ws_trades.cell(row=row_idx, column=5, value=trade.get("entry_price", 0))
                ws_trades.cell(row=row_idx, column=6, value=trade.get("exit_time", ""))
                ws_trades.cell(row=row_idx, column=7, value=trade.get("exit_price", ""))
                ws_trades.cell(row=row_idx, column=8, value=trade.get("quantity", 0))
                ws_trades.cell(row=row_idx, column=9, value=trade.get("pnl", 0))
                ws_trades.cell(row=row_idx, column=10, value=trade.get("status", ""))

            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                ws_trades.column_dimensions[col].width = 15

        # ==========================================
        # Symbol Performance Sheet
        # ==========================================
        if data.symbol_performance:
            ws_symbols = wb.create_sheet("By Symbol")

            headers = ["Symbol", "Trades", "Winners", "Losers", "P&L", "Win Rate"]
            for col, header in enumerate(headers, start=1):
                cell = ws_symbols.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row_idx, (symbol, perf) in enumerate(data.symbol_performance.items(), start=2):
                win_rate = perf["winners"] / perf["trades"] if perf["trades"] > 0 else 0

                ws_symbols.cell(row=row_idx, column=1, value=symbol)
                ws_symbols.cell(row=row_idx, column=2, value=perf["trades"])
                ws_symbols.cell(row=row_idx, column=3, value=perf["winners"])
                ws_symbols.cell(row=row_idx, column=4, value=perf["losers"])
                ws_symbols.cell(row=row_idx, column=5, value=perf["pnl"])
                ws_symbols.cell(row=row_idx, column=6, value=f"{win_rate*100:.1f}%")

        # Save workbook
        wb.save(filepath)

        logger.info(f"Excel report generated: {filepath}")
        return filepath

    def generate_html(
        self,
        start_date: date,
        end_date: date,
        filename: str = None
    ) -> str:
        """Generate HTML report"""
        data = self.collect_data(start_date, end_date)

        if not filename:
            filename = f"trading_report_{start_date}_{end_date}.html"

        filepath = os.path.join(self.config.output_dir, filename)

        pnl_color = "#43A047" if data.total_pnl >= 0 else "#E53935"

        html = f"""
<!DOCTYPE html>
<html>
<head>
    <title>{self.config.title}</title>
    <meta charset="utf-8">
    <script src="https://cdn.tailwindcss.com"></script>
</head>
<body class="bg-gray-100 p-8">
    <div class="max-w-4xl mx-auto">
        <h1 class="text-3xl font-bold text-blue-600 mb-2">{self.config.title}</h1>
        <p class="text-gray-600 mb-8">{start_date} to {end_date}</p>
        
        <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mb-8">
            <div class="bg-white p-4 rounded-lg shadow">
                <h3 class="text-gray-500 text-sm">Starting Equity</h3>
                <p class="text-xl font-bold">${data.starting_equity:,.2f}</p>
            </div>
            <div class="bg-white p-4 rounded-lg shadow">
                <h3 class="text-gray-500 text-sm">Ending Equity</h3>
                <p class="text-xl font-bold">${data.ending_equity:,.2f}</p>
            </div>
            <div class="bg-white p-4 rounded-lg shadow">
                <h3 class="text-gray-500 text-sm">Total P&L</h3>
                <p class="text-xl font-bold" style="color: {pnl_color}">${data.total_pnl:+,.2f}</p>
            </div>
            <div class="bg-white p-4 rounded-lg shadow">
                <h3 class="text-gray-500 text-sm">Return</h3>
                <p class="text-xl font-bold" style="color: {pnl_color}">{data.total_pnl_pct*100:+.2f}%</p>
            </div>
        </div>
        
        <div class="bg-white p-6 rounded-lg shadow mb-8">
            <h2 class="text-xl font-bold mb-4">Performance by Symbol</h2>
            <table class="w-full">
                <thead>
                    <tr class="bg-blue-500 text-white">
                        <th class="p-2 text-left">Symbol</th>
                        <th class="p-2 text-right">Trades</th>
                        <th class="p-2 text-right">P&L</th>
                        <th class="p-2 text-right">Win Rate</th>
                    </tr>
                </thead>
                <tbody>
"""

        for symbol, perf in data.symbol_performance.items():
            win_rate = perf["winners"] / perf["trades"] * 100 if perf["trades"] > 0 else 0
            row_color = "#43A047" if perf["pnl"] >= 0 else "#E53935"
            html += f"""
                    <tr class="border-b">
                        <td class="p-2">{symbol}</td>
                        <td class="p-2 text-right">{perf["trades"]}</td>
                        <td class="p-2 text-right" style="color: {row_color}">${perf["pnl"]:+,.2f}</td>
                        <td class="p-2 text-right">{win_rate:.1f}%</td>
                    </tr>
"""

        html += f"""
                </tbody>
            </table>
        </div>
        
        <p class="text-gray-500 text-sm">
            Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} by {self.config.author}
        </p>
    </div>
</body>
</html>
"""

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)

        logger.info(f"HTML report generated: {filepath}")
        return filepath


class ScheduledReportSender:
    """
    Scheduled report generation and sending.
    """

    def __init__(
        self,
        generator: ReportGenerator,
        feishu_webhook: str = None,
        email_config: Dict = None
    ):
        self.generator = generator
        self.feishu_webhook = feishu_webhook
        self.email_config = email_config
        self._running = False

    def send_daily_report(self):
        """Generate and send daily report"""
        today = date.today()

        try:
            # Generate PDF
            pdf_path = self.generator.generate_pdf(today, today)

            # Send via Feishu
            if self.feishu_webhook:
                self._send_feishu_notification(pdf_path, today)

            # Send via email
            if self.email_config:
                self._send_email(pdf_path, today)

            logger.info(f"Daily report sent for {today}")

        except Exception as e:
            logger.error(f"Failed to send daily report: {e}")

    def send_weekly_report(self):
        """Generate and send weekly report"""
        today = date.today()
        week_start = today - timedelta(days=today.weekday())

        try:
            pdf_path = self.generator.generate_pdf(week_start, today)
            excel_path = self.generator.generate_excel(week_start, today)

            if self.feishu_webhook:
                self._send_feishu_notification(pdf_path, today, "Weekly")

            logger.info(f"Weekly report sent for {week_start} to {today}")

        except Exception as e:
            logger.error(f"Failed to send weekly report: {e}")

    def _send_feishu_notification(self, filepath: str, report_date: date, report_type: str = "Daily"):
        """Send Feishu notification about report"""
        import requests

        if not self.feishu_webhook:
            return

        payload = {
            "msg_type": "interactive",
            "card": {
                "header": {
                    "title": {
                        "tag": "plain_text",
                        "content": f"📊 {report_type} Trading Report - {report_date}"
                    },
                    "template": "blue"
                },
                "elements": [
                    {
                        "tag": "div",
                        "text": {
                            "tag": "lark_md",
                            "content": f"**{report_type} trading report** has been generated.\n\nFile: `{filepath}`"
                        }
                    }
                ]
            }
        }

        try:
            requests.post(self.feishu_webhook, json=payload, timeout=10)
        except Exception as e:
            logger.error(f"Failed to send Feishu notification: {e}")

    def _send_email(self, filepath: str, report_date: date):
        """Send email with report attachment"""
        # Email sending implementation
        pass
